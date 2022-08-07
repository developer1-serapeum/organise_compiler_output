import re
from openpyxl import Workbook

BUILD_PATH_PATTERN = "\"D:/a/workspace/1ST201744/"

"""
The goal of this script is to get build log as text files and extract
compiler warnings. The complier warnings will be written to an excel sheet
with the columns (Warning id, file name, line number, and warning tile)  
"""

def convertToLine(buildLogRaw, buildLogOrganised):
    """
    Reads the build output text file where the waring is written across
    multiple lines. Then it writes a file where each line has one warning
    @param buildLogRaw: The output of the compilation process from the SDK
    @param buildLogOrganised: conatins al compiler warnings, where each 
        line has exactly one warning
    """

    file_in = buildLogRaw
    file_out = buildLogOrganised

    pattern = re.compile(BUILD_PATH_PATTERN)
    pattern2 = re.compile(r"^\s*\^")

    collect_lines = False
    new_lines = []
    conactenated_line = ""

    with open(file_in, 'r') as reader:
        match_count = 0
        line = reader.readline()
        while line != '':  # The EOF char is an empty string
            match = re.search(pattern, line)
            if match:
                #print("start of a warning sentence is found !")
                collect_lines = True
                match_count += 1

            match2 = re.search(pattern2, line)
            if match2:
                #print("The end of a warning sentence is found !")
                collect_lines = False
                new_lines.append(conactenated_line)
                #print(conactenated_line)
                conactenated_line = ""

            if(collect_lines):
                conactenated_line += line.replace("\n", " ")

            # read next line
            line = reader.readline()
        
    print("Found compiler warnings count ({})".format(match_count))

    # Write to the file where each line has exactly one warning
    with open(file_out, 'w') as f:
        for line in new_lines:
            f.write("%s\n" % line)


def extractWarningId(compilerInOrganisedFile):
    """
    Reads the organised compiler text file and returns an array which contains
    all warning ids
    """

    pattern = re.compile(r"warning\s(#.+-.):")

    match_count = 0
    found_instances = []
    for line in open(compilerInOrganisedFile):
        for match in re.finditer(pattern, line):
            found_instances.append(match.group(1))
            match_count += 1

    print("Warning ids count = ({}) times".format(match_count))
    return found_instances

def extractLineNumber(compilerInOrganisedFile):

    pattern = re.compile(r",\sline\s([0-9]+):")

    match_count = 0
    found_instances = []
    for line in open(compilerInOrganisedFile):
        for match in re.finditer(pattern, line):
            found_instances.append(match.group(1))
            match_count += 1

    print("Line numbers count = ({}) times".format(match_count))
    return found_instances

def extractFileName(compilerInOrganisedFile):


    pattern = re.compile(BUILD_PATH_PATTERN+"(.*)\",")

    match_count = 0
    found_instances = []
    for line in open(compilerInOrganisedFile):
        for match in re.finditer(pattern, line):
            found_instances.append(match.group(1))
            match_count += 1

    print("File names count = ({}) times".format(match_count))
    return found_instances

def extractWarningTitle(compilerInOrganisedFile):
    """
    """
    
    pattern = re.compile("warning #.+-.:\s+(.*)")

    match_count = 0
    found_instances = []
    for line in open(compilerInOrganisedFile):
        for match in re.finditer(pattern, line):
            found_instances.append(match.group(1))
            match_count += 1

    print("Warning titles count = ({}) times".format(match_count))

    return found_instances

def writeToExcel(warning_ids, lineNumbers, file_names, warning_titles):
    """
    Takes 4 input arrarys and writes them as columns in an excell sheet
    """
    workbook = Workbook()
    sheet = workbook.active
    #sheet["A1"] = "Hello World !"
    for row in range(1, len(lineNumbers)+1):
        sheet.cell(row=row, column=1).value = warning_ids[row-1]
        sheet.cell(row=row, column=2).value = file_names[row-1]
        sheet.cell(row=row, column=3).value = lineNumbers[row-1]
        sheet.cell(row=row, column=4).value = warning_titles[row-1]
    
    workbook.save(filename="compiler_out2.xlsx")


def findPattern():
    """
    [For debug only] Finds a desired pattern within a text file
    and writes it to an output file
    """
    file_in = 'compiler_in.txt'
    file_out = 'compiler_warning_ids.txt'
    pattern = re.compile(r"warning\s(#.+-.):")
    
    match_count = 0
    found_instances = []
    for i, line in enumerate(open(file_in)):
        for match in re.finditer(pattern, line):
            print ('Found on line %s: %s' % (i+1, match.group()) )
            found_instances.append(match.group(1))
            match_count += 1

    print("The pattern was found ({}) times".format(match_count))

    with open(file_out, 'w') as f:
        for line in found_instances:
            f.write("%s\n" % line)


def print_rows():
    """
    [For debug only] It makes it easier to print all of your spreadsheet values by
    just calling print_rows().
    """
    workbook = Workbook()
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        print(row)


###############################################################################
### The Main Program                                                        ### 
###############################################################################

if __name__ == "__main__":
    
    
    
    buildLogRaw = "compiler_in_example.txt"
    buildLogOrganised = buildLogRaw[:-4]+"_organised.txt"
    convertToLine(buildLogRaw, buildLogOrganised)
    
    #findPattern()
    warning_ids = extractWarningId(buildLogOrganised)
    line_numbers = extractLineNumber(buildLogOrganised)
    file_names = extractFileName(buildLogOrganised)
    warning_titles = extractWarningTitle(buildLogOrganised)
    
    writeToExcel(warning_ids, line_numbers, file_names, warning_titles)
    print ('The excel sheet {"compiler_out.xlsx"} was written successfully...')